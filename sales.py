import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font
files=glob.glob('input_file/*.csv')
for file in files:
    print(file)

for file in files:
    df=pd.read_csv(file)
    print(df.head())

df_list=[]

for file in files:
    try:
     df=pd.read_csv(file)
     df_list.append(df)
    except Exception as e:
        print(f'Error reading {file}:{e}')
combined=pd.concat(df_list)#concat:it sticks data tables together
print(combined)

combined.columns=combined.columns.str.strip().str.lower()
print(combined.columns)

combined.rename(columns={'product_name':'product','qty':'quantity','price_usd':'price'},inplace=True)
#inplace=True:modifies the main dataframe directly
print(combined.columns)

combined=combined.T.groupby(level=0).first().T


combined['product']=combined['product'].str.strip().str.capitalize()
print(combined)

print(combined.isnull().sum())

combined['quantity']=combined['quantity'].fillna(0)
combined['price']=combined['price'].fillna(0)
print(combined.isnull().sum())

combined.drop_duplicates(inplace=True)
print(combined.shape)



wb=load_workbook('output_file/cleaned_report.xlsx')
ws=wb.active

for cell in ws[1]:
    cell.font=Font(bold=True)

for col in ws.columns:
    max_length=0
    col_letter=col[0].column_letter

    for cell in col:
        if cell.value:
            max_length=max(max_length,len(str(cell.value)))

    ws.column_dimensions[col_letter].width=max_length+2
wb.save('output_file/cleaned_report.xlsx')

summary=combined.groupby('product').agg({'quantity':'sum','price':'mean'}).sort_values(by='quantity',ascending=False)

with pd.ExcelWriter('output_file/cleaned_report.xlsx', engine='openpyxl')as writer:
    combined.to_excel(writer,sheet_name='Cleaned Data',index=False)
    summary.to_excel(writer,sheet_name='Summary')

ws1=wb['Summary']
for cell in ws1[1]:
    cell.font=Font(bold=True)

for col in ws1.columns:
    max_length=0
    col_letter=col[0].column_letter
    for cell in col:
        if cell.value:
            max_length=max(max_length,len(str(cell.value)))

    ws1.column_dimensions[col_letter].width=max_length+2




ws.freeze_panes='A2'
ws1.freeze_panes='A2'
wb.save('output_file/cleaned_report.xlsx')
